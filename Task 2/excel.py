import os
import pandas as pd
from natsort import natsorted
from openpyxl import load_workbook

dirName = './pozyxAPI_dane_pomiarowe'


def parse_learn_data():
    data = pd.DataFrame()
    for filename in natsorted(os.listdir(dirName)):
        if filename.endswith(".xlsx"):
            df = pd.read_excel(f"{dirName}/{filename}")
            df = df[['0/timestamp', 't', 'no', 'measurement x', 'measurement y', 'reference x', 'reference y']]
            data = data.append(df, ignore_index=True)
    df = pd.read_excel("./pozyxAPI_only_localization_dane_testowe_i_dystrybuanta.xlsx")
    df = df[['0/timestamp', 't', 'no', 'measurement x', 'measurement y', 'reference x', 'reference y']]
    data = data.append(df, ignore_index=True)
    data.to_csv("./dataset.csv")


def add_results_to_main_excel(predict_test):
    df_to_export = pd.DataFrame(predict_test)
    with pd.ExcelWriter('./tmp.xlsx') as writer:
        df_to_export.to_excel(writer, sheet_name='Results', index=False)
    export_workbook = load_workbook(filename='./tmp.xlsx')
    export_sheet = export_workbook.active
    target_workbook = load_workbook(filename='./pozyxAPI_only_localization_dane_testowe_i_dystrybuanta.xlsx')
    target_sheet = target_workbook.active
    export_values = []
    for value in export_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        export_values.append(value)
    os.remove('tmp.xlsx')
    for it in range(1, export_values.__len__()+1):
        target_sheet[f'O{it + 1}'] = export_values[it-1][0]
        target_sheet[f'P{it + 1}'] = export_values[it-1][1]
    target_workbook.save('./Results.xlsx')
